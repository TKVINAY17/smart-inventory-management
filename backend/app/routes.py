from datetime import datetime

from fastapi import APIRouter, Depends, Header, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import load_workbook, Workbook

from app.database import get_db
from app.models import (
    User,
    Product,
    Sale,
    Supplier,
    PurchaseOrder,
    PurchaseOrderItem,
    Customer,
)
from app.schemas import (
    UserCreate,
    LoginRequest,
    ProductCreate,
    ProductUpdate,
    ProductResponse,
    SupplierCreate,
    SupplierResponse,
    PurchaseOrderCreate,
    PurchaseOrderResponse,
    Customer as CustomerSchema,
    CustomerCreate,
    CustomerUpdate,
)
from app.auth import (
    hash_password,
    verify_password,
    create_access_token,
    verify_token,
)

import shutil
import os
import io
import qrcode


router = APIRouter()


@router.get("/me")
def get_current_user(
    authorization: str = Header(None),
    db: Session = Depends(get_db),
):
    if authorization is None:
        raise HTTPException(status_code=401, detail="Authorization header missing")

    if not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Invalid token format")

    token = authorization.split(" ", 1)[1]
    email = verify_token(token)

    if email is None:
        raise HTTPException(status_code=401, detail="Invalid or expired token")

    user = db.query(User).filter(User.email == email).first()

    if user is None:
        raise HTTPException(status_code=404, detail="User not found")

    return {
        "id": user.id,
        "full_name": user.full_name,
        "email": user.email,
        "role": user.role,
    }


@router.post("/register")
def register(user: UserCreate, db: Session = Depends(get_db)):
    existing_user = db.query(User).filter(User.email == user.email).first()

    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")

    new_user = User(
        full_name=user.full_name,
        email=user.email,
        password=hash_password(user.password),
    )

    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    return {"message": "User registered successfully!"}


@router.post("/login")
def login(user: LoginRequest, db: Session = Depends(get_db)):
    db_user = db.query(User).filter(User.email == user.email).first()

    if db_user is None:
        raise HTTPException(status_code=401, detail="Invalid email or password")

    if not verify_password(user.password, db_user.password):
        raise HTTPException(status_code=401, detail="Invalid email or password")

    access_token = create_access_token(data={"sub": db_user.email})

    return {
        "access_token": access_token,
        "token_type": "bearer",
    }


@router.post("/upload-image")
def upload_image(file: UploadFile = File(...)):
    upload_folder = "uploads"
    os.makedirs(upload_folder, exist_ok=True)

    safe_filename = os.path.basename(file.filename)
    file_path = os.path.join(upload_folder, safe_filename)

    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    return {
        "filename": safe_filename,
        "url": f"/uploads/{safe_filename}",
    }


@router.post("/products", response_model=ProductResponse)
def add_product(product: ProductCreate, db: Session = Depends(get_db)):
    new_product = Product(
        name=product.name,
        description=product.description,
        category=product.category,
        price=product.price,
        quantity=product.quantity,
        image=product.image,
    )

    db.add(new_product)
    db.commit()
    db.refresh(new_product)
    return new_product


@router.get("/products", response_model=list[ProductResponse])
def get_products(db: Session = Depends(get_db)):
    return db.query(Product).all()


@router.get("/dashboard")
def dashboard(db: Session = Depends(get_db)):
    products = db.query(Product).all()
    sales = db.query(Sale).all()

    total_products = len(products)
    inventory_value = sum(product.price * product.quantity for product in products)
    low_stock = len([product for product in products if product.quantity < 10])
    total_sales = len(sales)
    revenue = sum(sale.total_price for sale in sales)

    return {
        "totalProducts": total_products,
        "inventoryValue": inventory_value,
        "lowStock": low_stock,
        "totalSales": total_sales,
        "revenue": revenue,
    }


@router.get("/dashboard/charts")
def dashboard_charts(db: Session = Depends(get_db)):
    products = db.query(Product).all()

    labels = []
    stock = []
    category_count = {}

    for product in products:
        labels.append(product.name)
        stock.append(product.quantity)
        category_count[product.category] = category_count.get(product.category, 0) + 1

    return {
        "labels": labels,
        "stock": stock,
        "categories": category_count,
    }


@router.put("/products/{product_id}", response_model=ProductResponse)
def update_product(
    product_id: int,
    product: ProductUpdate,
    db: Session = Depends(get_db),
):
    db_product = db.query(Product).filter(Product.id == product_id).first()

    if db_product is None:
        raise HTTPException(status_code=404, detail="Product not found")

    db_product.name = product.name
    db_product.description = product.description
    db_product.category = product.category
    db_product.price = product.price
    db_product.quantity = product.quantity
    db_product.image = product.image

    db.commit()
    db.refresh(db_product)
    return db_product


@router.put("/products/{product_id}/restock", response_model=ProductResponse)
def restock_product(
    product_id: int,
    quantity: int,
    db: Session = Depends(get_db),
):
    if quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be greater than 0")

    db_product = db.query(Product).filter(Product.id == product_id).first()

    if db_product is None:
        raise HTTPException(status_code=404, detail="Product not found")

    db_product.quantity += quantity
    db.commit()
    db.refresh(db_product)
    return db_product


@router.delete("/products/{product_id}")
def delete_product(
    product_id: int,
    db: Session = Depends(get_db),
):
    product = db.query(Product).filter(Product.id == product_id).first()

    if product is None:
        raise HTTPException(status_code=404, detail="Product not found")

    db.delete(product)
    db.commit()
    return {"message": "Product deleted successfully"}


@router.post("/import-products")
def import_products(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    try:
        workbook = load_workbook(file.file)
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Invalid Excel file. Please upload a valid .xlsx file.",
        )

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    imported_count = 0
    errors = []

    for index, row in enumerate(rows[1:], start=2):
        try:
            name = row[0]
            description = row[1]
            category = row[2]
            price = row[3]
            quantity = row[4]

            if not name:
                errors.append(f"Row {index}: missing name, skipped")
                continue

            new_product = Product(
                name=name,
                description=description,
                category=category,
                price=float(price) if price is not None else 0.0,
                quantity=int(quantity) if quantity is not None else 0,
            )

            db.add(new_product)
            imported_count += 1

        except (ValueError, TypeError, IndexError) as e:
            errors.append(f"Row {index}: {str(e)}")

    db.commit()

    return {
        "message": f"Imported {imported_count} products successfully!",
        "errors": errors,
    }


@router.get("/export-products")
def export_products(db: Session = Depends(get_db)):
    products = db.query(Product).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append(["ID", "Name", "Description", "Category", "Price", "Quantity"])

    for product in products:
        sheet.append([
            product.id,
            product.name,
            product.description,
            product.category,
            product.price,
            product.quantity,
        ])

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)

    headers = {"Content-Disposition": "attachment; filename=products.xlsx"}

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/products/{product_id}/qrcode")
def generate_qrcode(
    product_id: int,
    db: Session = Depends(get_db),
):
    product = db.query(Product).filter(Product.id == product_id).first()

    if product is None:
        raise HTTPException(status_code=404, detail="Product not found")

    qr_data = f"""
📦 SMART INVENTORY MANAGEMENT

────────────────────────

🆔 Product ID : {product.id}

📱 Product Name : {product.name}

📝 Description : {product.description}

📂 Category : {product.category}

💰 Price : ₹{product.price}

📦 Stock : {product.quantity}

✅ Status : {"In Stock" if product.quantity > 10 else "Low Stock" if product.quantity > 0 else "Out of Stock"}

────────────────────────
Generated by Smart Inventory Management
"""

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=4,
    )

    qr.add_data(qr_data)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")

    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)

    return StreamingResponse(buffer, media_type="image/png")


@router.post("/suppliers", response_model=SupplierResponse)
def add_supplier(
    supplier: SupplierCreate,
    db: Session = Depends(get_db),
):
    new_supplier = Supplier(
        company_name=supplier.company_name,
        contact_person=supplier.contact_person,
        email=supplier.email,
        phone=supplier.phone,
        address=supplier.address,
        gst_number=supplier.gst_number,
    )

    db.add(new_supplier)
    db.commit()
    db.refresh(new_supplier)

    return new_supplier


@router.get("/suppliers", response_model=list[SupplierResponse])
def get_suppliers(db: Session = Depends(get_db)):
    return db.query(Supplier).all()


@router.put("/suppliers/{supplier_id}", response_model=SupplierResponse)
def update_supplier(
    supplier_id: int,
    supplier: SupplierCreate,
    db: Session = Depends(get_db),
):
    db_supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()

    if db_supplier is None:
        raise HTTPException(status_code=404, detail="Supplier not found")

    db_supplier.company_name = supplier.company_name
    db_supplier.contact_person = supplier.contact_person
    db_supplier.email = supplier.email
    db_supplier.phone = supplier.phone
    db_supplier.address = supplier.address
    db_supplier.gst_number = supplier.gst_number

    db.commit()
    db.refresh(db_supplier)

    return db_supplier


@router.delete("/suppliers/{supplier_id}")
def delete_supplier(
    supplier_id: int,
    db: Session = Depends(get_db),
):
    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()

    if supplier is None:
        raise HTTPException(status_code=404, detail="Supplier not found")

    db.delete(supplier)
    db.commit()

    return {"message": "Supplier deleted successfully"}


# ==========================================================
# PURCHASE ORDERS
# ==========================================================

@router.get(
    "/purchase-orders",
    response_model=list[PurchaseOrderResponse],
)
def get_purchase_orders(
    db: Session = Depends(get_db),
):
    return (
        db.query(PurchaseOrder)
        .order_by(PurchaseOrder.id.desc())
        .all()
    )


@router.post(
    "/purchase-orders",
    response_model=PurchaseOrderResponse,
)
def create_purchase_order(
    purchase_order: PurchaseOrderCreate,
    db: Session = Depends(get_db),
):

    last_order = (
        db.query(PurchaseOrder)
        .order_by(PurchaseOrder.id.desc())
        .first()
    )

    next_number = 1

    if last_order:
        next_number = last_order.id + 1

    po_number = f"PO-{next_number:04d}"

    total = 0

    new_order = PurchaseOrder(
        po_number=po_number,
        supplier_id=purchase_order.supplier_id,
        supplier_name=purchase_order.supplier_name,
        status="Pending",
    )

    db.add(new_order)
    db.commit()
    db.refresh(new_order)

    for item in purchase_order.items:

        item_total = item.quantity * item.unit_price

        total += item_total

        db_item = PurchaseOrderItem(
            purchase_order_id=new_order.id,
            product_id=item.product_id,
            product_name=item.product_name,
            quantity=item.quantity,
            unit_price=item.unit_price,
            total_price=item_total,
        )

        db.add(db_item)

    new_order.total_amount = total

    db.commit()
    db.refresh(new_order)

    return new_order


@router.delete("/purchase-orders/{order_id}")
def delete_purchase_order(
    order_id: int,
    db: Session = Depends(get_db),
):

    order = (
        db.query(PurchaseOrder)
        .filter(PurchaseOrder.id == order_id)
        .first()
    )

    if order is None:
        raise HTTPException(
            status_code=404,
            detail="Purchase Order not found",
        )

    db.query(PurchaseOrderItem).filter(
        PurchaseOrderItem.purchase_order_id == order.id
    ).delete()

    db.delete(order)

    db.commit()

    return {
        "message": "Purchase Order deleted successfully"
    }


@router.post("/purchase-orders/{po_id}/receive")
def receive_purchase_order(po_id: int, db: Session = Depends(get_db)):

    po = (
        db.query(PurchaseOrder)
        .filter(PurchaseOrder.id == po_id)
        .first()
    )

    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")

    if po.status == "Received":
        raise HTTPException(
            status_code=400,
            detail="Purchase Order already received",
        )

    for item in po.items:

        product = (
            db.query(Product)
            .filter(Product.id == item.product_id)
            .first()
        )

        if product:
            product.quantity += item.quantity

    po.status = "Received"
    po.received_date = datetime.utcnow()

    db.commit()

    return {
        "message": "Goods received successfully"
    }


# ==========================================================
# CUSTOMERS
# ==========================================================

@router.get("/customers", response_model=list[CustomerSchema])
def get_customers(db: Session = Depends(get_db)):
    return db.query(Customer).all()


@router.post("/customers", response_model=CustomerSchema)
def create_customer(customer: CustomerCreate, db: Session = Depends(get_db)):
    new_customer = Customer(**customer.model_dump())
    db.add(new_customer)
    db.commit()
    db.refresh(new_customer)
    return new_customer


@router.put("/customers/{customer_id}", response_model=CustomerSchema)
def update_customer(
    customer_id: int,
    customer: CustomerUpdate,
    db: Session = Depends(get_db),
):
    db_customer = db.query(Customer).filter(Customer.id == customer_id).first()

    if not db_customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    for key, value in customer.model_dump().items():
        setattr(db_customer, key, value)

    db.commit()
    db.refresh(db_customer)

    return db_customer


@router.delete("/customers/{customer_id}")
def delete_customer(customer_id: int, db: Session = Depends(get_db)):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()

    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    db.delete(customer)
    db.commit()

    return {"message": "Customer deleted successfully"}